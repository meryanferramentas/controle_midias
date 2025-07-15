from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

entrada = './planilha/planilha_com_abas.xlsx'
saida = './planilha/planilha_final.xlsx'

wb = load_workbook(entrada)

colunas_para_mesclar = ['Fornecedor', 'Descrição', 'SKU Pai']

for ws in wb.worksheets:
    print(f"Mesclando na aba: {ws.title}")
    header = {cell.value: idx + 1 for idx, cell in enumerate(ws[1])}

    for col_nome in colunas_para_mesclar:
        col_idx = header.get(col_nome)
        if not col_idx:
            continue
        start = 2
        last_val = ws.cell(row=start, column=col_idx).value

        for row in range(3, ws.max_row + 2):
            val = ws.cell(
                row, column=col_idx).value if row <= ws.max_row else None
            if val != last_val:
                if row - 1 > start:
                    for r in range(start, row):
                        ws.cell(row=r, column=col_idx).value = last_val
                    ws.merge_cells(
                        f"{get_column_letter(col_idx)}{start}:{get_column_letter(col_idx)}{row - 1}")
                start = row
                last_val = val

wb.save(saida)
print(f"Planilha com abas mescladas salva com sucesso: {saida}")
